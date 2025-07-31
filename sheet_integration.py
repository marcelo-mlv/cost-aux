import pandas as pd
from anytree import Node, RenderTree
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

def load_BOM_df(filename):
    path = './' + filename
    df = pd.read_excel(path, sheet_name='BOM', header=5, dtype=str)
    return df

def preprocess(df):
    # Remove white rows (sum of a subsystem)
    df = df[df['Asm/Prt #'].notna()]
    return df

def create_hierarchy(df):
    # Root: Parent of all subsystems
    BOM_root = Node("BOM", parent=None)
    subsystem_nodes = {}
    assembly_nodes = {}

    for _, row in df.iterrows():
        subsystem = row['Area of Commodity']
        asm_prt = row['Asm/Prt #']
        component = row['Component']

        # Create subsystem node if it doesn't exist already
        if subsystem not in subsystem_nodes:
            subsystem_nodes[subsystem] = Node(subsystem, parent=BOM_root)

        # Create assembly node as child of subsystem node
        if asm_prt[0] == 'A': # Assembly
            assembly_nodes[component] = Node(component, parent=subsystem_nodes[subsystem])
        # Part (under an assembly)
        elif assembly_nodes:
            Node(component, parent=list(assembly_nodes.values())[-1])
        # If no assembly node exists, raise an error
        else:
            raise ValueError(f"Component {component} does not have a parent assembly. The first BOM row is probably a Part.")
    return BOM_root

def get_BOM_df(filename):
    BOM_raw = load_BOM_df(filename)
    BOM_data = preprocess(BOM_raw)
    return BOM_data

def get_BOM_tree(filename):
    BOM_data = get_BOM_df(filename)
    BOM_root = create_hierarchy(BOM_data)
    return BOM_root

def save_tree_to_txt(BOM_root):
    with open("bom_tree.txt", "w", encoding="utf-8") as f:
        for pre, _, node in RenderTree(BOM_root):
            f.write(f"{pre}{node.name}\n")

def get_assembly_parts_mapping(filename):
    """
    Retorna um dicionário mapeando cada assembly para suas parts filhas.
    Baseado na estrutura do BOM onde assemblies começam com 'A' e parts são seus filhos diretos.
    
    Returns:
        dict: {assembly_name: [list_of_part_names]}
    """
    df = get_BOM_df(filename)
    assembly_parts = {}
    current_assembly = None
    
    for _, row in df.iterrows():
        asm_prt = row['Asm/Prt #']
        component = row['Component']
        
        # Se é um assembly (começa com 'A')
        if asm_prt[0] == 'A':
            current_assembly = component
            assembly_parts[current_assembly] = []
        # Se é uma part e temos um assembly atual
        elif current_assembly is not None:
            assembly_parts[current_assembly].append(component)
    
    return assembly_parts

def create_hyperlinks_for_assemblies(filename):
    """
    Cria hyperlinks nas páginas de assembly para suas respectivas parts.
    Os hyperlinks são colocados nas células B9, B10, B11, etc.
    
    Args:
        filename (str): Nome do arquivo Excel
        
    Returns:
        dict: Relatório de hyperlinks criados
    """
    try:
        # Carregar o workbook com openpyxl
        path = './' + filename
        workbook = load_workbook(path)
        
        # Obter mapeamento de assemblies e parts
        assembly_parts = get_assembly_parts_mapping(filename)
        
        report = {
            'assemblies_processed': 0,
            'hyperlinks_created': 0,
            'missing_sheets': [],
            'details': {}
        }
        
        # Para cada assembly, criar hyperlinks
        for assembly_name, parts_list in assembly_parts.items():
            assembly_info = {
                'parts_count': len(parts_list),
                'hyperlinks_created': 0,
                'missing_parts_sheets': []
            }
            
            # Verificar se existe uma worksheet com o nome do assembly
            if assembly_name in workbook.sheetnames:
                worksheet = workbook[assembly_name]
                
                # Começar na célula B9
                start_row = 9
                
                for i, part_name in enumerate(parts_list):
                    item_order_address = f"A{start_row + i}"
                    cellpart_address = f"B{start_row + i}"
                    partcost_address = f"C{start_row + i}"
                    quantity_address = f"D{start_row + i}"
                    subtotal_address = f"E{start_row + i}"
                    
                    # Verificar se existe uma worksheet com o nome da part
                    if part_name in workbook.sheetnames:
                        worksheet[item_order_address].value = (i+1) * 10

                        # Criar hyperlink para a página da part (sheets)
                        worksheet[cellpart_address].value = f'=HYPERLINK("#{part_name}!A1", "{part_name}")'
                        
                        # Criar hyperlink para a página da part (excel)
                        # worksheet[cellpart_address].hyperlink = f"#{part_name}!A1"
                        # worksheet[cellpart_address].value = part_name
                        
                        worksheet[cellpart_address].style = "Hyperlink"
                        assembly_info['hyperlinks_created'] += 1
                        report['hyperlinks_created'] += 1

                        # Adicionar fórmulas para partcost e quantity
                        worksheet[partcost_address].value = f"=INDIRECT(\"\'\"&${cellpart_address}&\"\'!N1\")"
                        worksheet[quantity_address].value = f"=INDIRECT(\"\'\"&${cellpart_address}&\"\'!N2\")"
                        worksheet[subtotal_address].value = f"={partcost_address}*{quantity_address}"
                    else:
                        # Se não existe página da part, apenas colocar o nome
                        worksheet[cellpart_address].value = part_name
                        assembly_info['missing_parts_sheets'].append(part_name)
                
                report['assemblies_processed'] += 1
            else:
                report['missing_sheets'].append(assembly_name)
            
            report['details'][assembly_name] = assembly_info
        
        # Salvar o arquivo
        workbook.save(path)
        workbook.close()
        
        return report
        
    except Exception as e:
        return {'error': str(e)}

def print_hyperlink_report(report):
    """
    Imprime um relatório detalhado dos hyperlinks criados.
    
    Args:
        report (dict): Relatório retornado por create_hyperlinks_for_assemblies
    """
    if 'error' in report:
        print(f"Erro ao criar hyperlinks: {report['error']}")
        return
    
    print("=== RELATÓRIO DE CRIAÇÃO DE HYPERLINKS ===")
    print(f"Assemblies processados: {report['assemblies_processed']}")
    print(f"Total de hyperlinks criados: {report['hyperlinks_created']}")
    
    if report['missing_sheets']:
        print(f"\nPáginas de assembly não encontradas: {len(report['missing_sheets'])}")
        for sheet in report['missing_sheets']:
            print(f"  - {sheet}")
    
    print("\nDetalhes por assembly:")
    for assembly, info in report['details'].items():
        print(f"\n{assembly}:")
        print(f"  Parts totais: {info['parts_count']}")
        print(f"  Hyperlinks criados: {info['hyperlinks_created']}")
        
        if info['missing_parts_sheets']:
            print(f"  Parts sem página própria: {len(info['missing_parts_sheets'])}")
            for part in info['missing_parts_sheets']:
                print(f"    - {part}")
    
    print("\n=== FIM DO RELATÓRIO ===")

